import httpx
import json
import html2text
import io
from typing import Optional, Dict, Any

from app.config import settings


class LLMAnalyzer:
    """Сервис для анализа писем через локальную LLM (Ollama)."""
    
    def __init__(self):
        self.api_url = settings.OLLAMA_API_URL
        self.model = settings.OLLAMA_MODEL
        self.timeout = 180  # секунд
        
        # Конвертер HTML в текст
        self.html_converter = html2text.HTML2Text()
        self.html_converter.ignore_links = False
        self.html_converter.ignore_images = True
        self.html_converter.body_width = 0
    
    async def analyze_email(
        self,
        subject: str,
        body_text: str,
        from_email: str,
        attachments_text: str = ""
    ) -> Dict[str, Any]:
        """
        Отправляет письмо в LLM для анализа.
        """
        # Ограничиваем текст, чтобы не превысить контекст модели
        body_preview = body_text[:2000] if len(body_text) > 2000 else body_text
        
        # Формируем блок с содержимым вложений
        attachments_block = ""
        if attachments_text:
            # Ограничиваем общий объем текста из вложений
            attachments_block = f"""
СОДЕРЖИМОЕ ВЛОЖЕНИЙ:
{attachments_text[:8000]}
"""
        
        prompt = f"""Ты - ассистент для классификации электронной почты и анализа тендерной документации.

КРИТИЧЕСКИ ВАЖНО:
- Отвечай СТРОГО на РУССКОМ языке
- Не используй другие языки
- Все поле "summary" должно быть на русском языке

Проанализируй следующее письмо и вложения (если есть).

ОТПРАВИТЕЛЬ: {from_email}
ТЕМА: {subject}

ТЕКСТ ПИСЬМА:
{body_preview}
{attachments_block}

ЗАДАЧА:
1. Определи категорию письма. Выбери ОДНУ из:
   - TENDER: письмо о тендере, закупке, аукционе, конкурсе (44-ФЗ, 223-ФЗ, ЭТП). 
     ВАЖНО: НЕ классифицируй как TENDER новостные рассылки о госзакупках, 
     дайджесты ЭТП с подборкой тендеров, уведомления ГИС ЕИС без конкретных данных.
   - SPAM: реклама, спам, маркетинговая рассылка, акции, скидки
   - GENERAL: обычная деловая переписка, уведомления, личные письма
   - EMPTY: письмо пустое, без содержания

2. Напиши краткое резюме (1-2 предложения ОБЯЗАТЕЛЬНО на русском языке).

3. Если это TENDER, извлеки из ТЕКСТА ПИСЬМА и/или ВЛОЖЕНИЙ ключевую информацию:
   - notice_number: номер извещения (например, "0123456789")
   - purchase_name: название закупки (полное название)
   - nmck: начальная максимальная цена контракта (с валютой)
   - deadline: срок подачи заявок (дата)
   
   Если данных нет ни в теле письма, ни во вложениях - верни null для соответствующего поля.

ОТВЕТЬ СТРОГО В ФОРМАТЕ JSON (без markdown, без пояснений, ТОЛЬКО на русском):
{{
  "category": "TENDER|SPAM|GENERAL|EMPTY",
  "summary": "краткое резюме на русском языке",
  "tender_details": {{
    "notice_number": "номер извещения или null",
    "purchase_name": "название закупки или null",
    "nmck": "цена или null",
    "deadline": "срок подачи или null"
  }}
}}

Если это не тендер, верни "tender_details": null."""

        payload = {
            "model": self.model,
            "prompt": prompt,
            "stream": False,
            "format": "json"
        }
        
        try:
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    self.api_url,
                    json=payload,
                    timeout=self.timeout
                )
                response.raise_for_status()
                
                result = response.json()
                raw_response = result.get("response", "{}")
                
                # Парсим JSON ответ от модели
                analysis = json.loads(raw_response)
                
                # Проверяем язык резюме (защита от китайского)
                summary = analysis.get("summary", "")
                if self._is_non_russian(summary):
                    analysis["category"] = "EMPTY"
                    analysis["summary"] = "(нерусскоязычный ответ модели)"
                
                return analysis
                
        except httpx.TimeoutException:
            print(f"  ✗ Таймаут запроса к Ollama ({self.timeout}с)")
            return self._error_result("Таймаут запроса к LLM")
        except httpx.RequestError as e:
            print(f"  ✗ Ошибка запроса к Ollama: {e}")
            return self._error_result(f"Ошибка подключения: {str(e)}")
        except json.JSONDecodeError as e:
            print(f"  ✗ Ошибка парсинга JSON от LLM: {e}")
            return self._error_result("Модель вернула невалидный JSON")
        except Exception as e:
            print(f"  ✗ Непредвиденная ошибка: {e}")
            return self._error_result(f"Непредвиденная ошибка: {str(e)}")
    
    def _is_non_russian(self, text: str) -> bool:
        """Проверяет, содержит ли текст много нерусских символов."""
        if not text:
            return False
        
        # Считаем долю кириллицы
        cyrillic_count = sum(1 for c in text if 'а' <= c.lower() <= 'я')
        total_letters = sum(1 for c in text if c.isalpha())
        
        if total_letters == 0:
            return False
        
        cyrillic_ratio = cyrillic_count / total_letters
        
        # Если кириллицы меньше 30% — скорее всего не русский
        return cyrillic_ratio < 0.3
    
    def _error_result(self, error_message: str) -> Dict[str, Any]:
        """Возвращает структуру ошибки."""
        return {
            "category": "ERROR",
            "summary": error_message,
            "tender_details": None
        }
    
    def extract_text_from_html(self, html_content: str) -> str:
        """Конвертирует HTML в чистый текст."""
        return self.html_converter.handle(html_content)
    
    def extract_text_from_attachment(self, attachment_info: dict, payload: bytes) -> str:
        """
        Извлекает текст из вложения различных форматов.
        
        Поддерживаемые форматы:
        - PDF (pypdf)
        - DOCX (python-docx)
        - DOC (простое извлечение)
        - XLSX (openpyxl)
        - XLS (xlrd)
        - RTF (простая очистка)
        - TXT, XML
        - ZIP (только список файлов)
        
        Args:
            attachment_info: Словарь с информацией о вложении (filename, content_type)
            payload: Содержимое файла в байтах
        
        Returns:
            Извлеченный текст или пустая строка
        """
        filename = attachment_info.get("filename", "").lower()
        content_type = attachment_info.get("content_type", "").lower()
        
        try:
            # PDF файлы
            if content_type == "application/pdf" or filename.endswith(".pdf"):
                return self._extract_from_pdf(payload)
            
            # DOCX файлы
            elif (content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" 
                  or filename.endswith(".docx")):
                return self._extract_from_docx(payload)
            
            # DOC файлы (старый формат Word)
            elif (content_type == "application/msword" 
                  or filename.endswith(".doc")):
                return self._extract_from_doc(payload, filename)
            
            # XLSX файлы (Excel)
            elif (content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
                  or filename.endswith(".xlsx")):
                return self._extract_from_xlsx(payload)
            
            # XLS файлы (старый формат Excel)
            elif (content_type == "application/vnd.ms-excel" 
                  or filename.endswith(".xls")):
                return self._extract_from_xls(payload)
            
            # RTF файлы
            elif content_type == "application/rtf" or filename.endswith(".rtf"):
                return self._extract_from_rtf(payload, filename)
            
            # TXT файлы
            elif content_type == "text/plain" or filename.endswith(".txt"):
                return self._extract_from_txt(payload, filename)
            
            # XML файлы
            elif content_type in ["application/xml", "text/xml"] or filename.endswith(".xml"):
                return self._extract_from_txt(payload, filename)
            
            # ZIP файлы (показываем только список содержимого)
            elif (content_type in ["application/zip", "application/x-zip-compressed"] 
                  or filename.endswith(".zip")):
                return self._extract_from_zip(payload, filename)
            
            else:
                return f"[Неподдерживаемый формат: {filename} ({content_type})]"
        
        except Exception as e:
            print(f"    ⚠ Ошибка извлечения текста из {filename}: {e}")
            return ""
    
    def _extract_from_pdf(self, payload: bytes) -> str:
        """Извлечение текста из PDF."""
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(payload))
        text = ""
        # Берем первые 10 страниц
        for page in reader.pages[:10]:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text[:5000]
    
    def _extract_from_docx(self, payload: bytes) -> str:
        """Извлечение текста из DOCX."""
        from docx import Document
        doc = Document(io.BytesIO(payload))
        text = "\n".join([p.text for p in doc.paragraphs[:100]])
        return text[:5000]
    
    def _extract_from_doc(self, payload: bytes, filename: str) -> str:
        """Извлечение текста из DOC (старый формат)."""
        try:
            # Попытка извлечь текст напрямую
            text = payload.decode('utf-8', errors='ignore')
            # Удаляем бинарные символы
            text = ''.join(c for c in text if c.isprintable() or c in '\n\r\t')
            # Если текст слишком короткий, возможно это бинарный формат
            if len(text.strip()) < 100:
                return f"[DOC файл: {filename}, размер: {len(payload)} байт - бинарный формат]"
            return text[:5000]
        except Exception:
            return f"[DOC файл: {filename}, размер: {len(payload)} байт]"
    
    def _extract_from_xlsx(self, payload: bytes) -> str:
        """Извлечение текста из XLSX."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        text = ""
        for sheet in wb.worksheets[:3]:  # Первые 3 листа
            text += f"\n=== Лист: {sheet.title} ===\n"
            row_count = 0
            for row in sheet.iter_rows(max_row=100, values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
                    row_count += 1
                    if row_count >= 50:  # Ограничение по строкам
                        break
        wb.close()
        return text[:5000]
    
    def _extract_from_xls(self, payload: bytes) -> str:
        """Извлечение текста из XLS (старый формат)."""
        from xlrd import open_workbook
        wb = open_workbook(file_contents=payload)
        text = ""
        for sheet in wb.sheets()[:3]:  # Первые 3 листа
            text += f"\n=== Лист: {sheet.name} ===\n"
            for row_idx in range(min(50, sheet.nrows)):  # Первые 50 строк
                row = sheet.row_values(row_idx)
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text[:5000]
    
    def _extract_from_rtf(self, payload: bytes, filename: str) -> str:
        """Извлечение текста из RTF."""
        try:
            text = payload.decode('utf-8', errors='ignore')
            # Простая очистка RTF-тегов
            import re
            text = re.sub(r'\\[a-z]+\d*\s?', '', text)
            text = re.sub(r'[{}]', '', text)
            return text[:5000]
        except Exception:
            return f"[RTF файл: {filename}]"
    
    def _extract_from_txt(self, payload: bytes, filename: str) -> str:
        """Извлечение текста из TXT/XML."""
        try:
            # Пробуем разные кодировки
            for encoding in ['utf-8', 'cp1251', 'cp866', 'koi8-r']:
                try:
                    text = payload.decode(encoding)
                    return text[:5000]
                except UnicodeDecodeError:
                    continue
            # Если не получилось, используем ignore
            text = payload.decode('utf-8', errors='ignore')
            return text[:5000]
        except Exception:
            return f"[Текстовый файл: {filename}]"
    
    def _extract_from_zip(self, payload: bytes, filename: str) -> str:
        """Извлечение списка файлов из ZIP."""
        import zipfile
        try:
            with zipfile.ZipFile(io.BytesIO(payload)) as zf:
                file_list = zf.namelist()
                text = f"[ZIP архив '{filename}' содержит {len(file_list)} файлов]:\n"
                for f in file_list[:20]:  # Первые 20 файлов
                    text += f"  - {f}\n"
                if len(file_list) > 20:
                    text += f"  ... и еще {len(file_list) - 20} файлов\n"
                
                # Пытаемся извлечь текст из текстовых файлов внутри архива
                for f in file_list[:5]:  # Первые 5 файлов
                    if f.lower().endswith(('.txt', '.xml', '.csv')):
                        try:
                            with zf.open(f) as zf_file:
                                content = zf_file.read()
                                inner_text = content.decode('utf-8', errors='ignore')[:1000]
                                text += f"\n--- Содержимое {f} ---\n{inner_text}\n"
                        except Exception:
                            pass
                return text
        except Exception as e:
            return f"[ZIP файл: {filename}, ошибка: {e}]"


# Функция для тестирования
async def test_llm_analyzer():
    """Тест работы LLM-анализатора."""
    analyzer = LLMAnalyzer()
    
    print("=" * 60)
    print("ТЕСТ LLM-АНАЛИЗАТОРА")
    print("=" * 60)
    print(f"Модель: {analyzer.model}")
    print(f"API URL: {analyzer.api_url}")
    print()
    
    # Тест 1: Тендерное письмо
    print("Тест 1: Тендерное письмо")
    result = await analyzer.analyze_email(
        subject="Извещение о проведении электронного аукциона №0123456789",
        body_text="Уважаемый участник! Извещаем вас о проведении электронного аукциона на поставку серверного оборудования. НМЦК: 2 500 000 руб. Срок подачи заявок: 05.07.2026 до 10:00 МСК.",
        from_email="info@sberbank-ast.ru"
    )
    print(f"  Категория: {result.get('category')}")
    print(f"  Резюме: {result.get('summary')}")
    print(f"  Детали тендера: {result.get('tender_details')}")
    print()
    
    # Тест 2: Спам
    print("Тест 2: Рекламное письмо")
    result = await analyzer.analyze_email(
        subject="Скидка 50% на все услуги!",
        body_text="Только сегодня! Успейте воспользоваться специальным предложением. Акция действует до конца месяца.",
        from_email="promo@spam-company.ru"
    )
    print(f"  Категория: {result.get('category')}")
    print(f"  Резюме: {result.get('summary')}")
    print()
    
    # Тест 3: Обычное письмо
    print("Тест 3: Обычное деловое письмо")
    result = await analyzer.analyze_email(
        subject="Re: Встреча в понедельник",
        body_text="Коллеги, добрый день! Подтверждаю встречу на понедельник в 14:00. Повестку отправлю позже.",
        from_email="colleague@company.ru"
    )
    print(f"  Категория: {result.get('category')}")
    print(f"  Резюме: {result.get('summary')}")
    print()
    
    print("=" * 60)
    print("Тест завершен")
    print("=" * 60)


if __name__ == "__main__":
    import asyncio
    asyncio.run(test_llm_analyzer())